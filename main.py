from openpyxl import Workbook, load_workbook
import itertools
from time import sleep


THREAD_TOTAL_WEEK_NUMBER = 20
PERSON_NUMBER = 5

# 1 person included
MIN_PERSON_NUMBER_WORKING_PER_WEEKEND = 1
# 2 persons included
MAX_PERSON_NUMBER_WORKING_PER_WEEKEND = 2

MIN_PERSON_NUMBER_WORKING_PER_WORKING_DAY = 1
MAX_WEEKLY_WORKING_HOURS = 48
ONE_DAY_WORKING_HOUR = 12
WORKING_DAY = 'J'
DAY_OFF = '-'

variant_list = []
valid_res = []


def detect_if_valid_number_of_persons_per_weekend(v) -> bool:
    if len(v[0])%7 != 0:
        raise ValueError("Error, thread length is not modulo %7, strange ! -> {}".format(len(v[0])))
    week_number = int(len(v[0])/7)
    for i in range(0, week_number):
        # person_working_per_week_end -> res
        res = 0
        for p in range(0, PERSON_NUMBER):
            if v[p][7*i+5] != v[p][7*i+6]:
                raise ValueError("Error, variant {}, we have detected a person working only one "
                                 "day in the weekend".format(v))
            # checking saturday (+5)
            if v[p][7*i+5] == WORKING_DAY:
                res = res+1
        if res < MIN_PERSON_NUMBER_WORKING_PER_WEEKEND or res > MAX_PERSON_NUMBER_WORKING_PER_WEEKEND:
            print("Variant {} is invalid because there are {} persons working on the {}th weekend".format(v, res, i+1))
            return False
    return True


def detect_if_one_person_per_working_day(v):
    week_number = int(len(v[0]) / 7)
    for i in range(0, week_number):
        for d in range(0, 5):
            res = 0
            for p in range(0, PERSON_NUMBER):
                if v[p][7*i+d] == WORKING_DAY:
                    res = res+1
                    break
            # print("number of persons working on day {} -> {}".format(i+1, res))
            if res < MIN_PERSON_NUMBER_WORKING_PER_WORKING_DAY:
                # print("No sufficient working person number on day {} on Variant {}".format(i+1, v))
                return False
    return True


def count_number_of_one_person_per_day_per_working_day(v) -> int:
    end_res = 0
    week_number = int(len(v[0])/7)
    for i in range(0, week_number):
        for d in range(0, 5):
            res = 0
            for p in range(0, PERSON_NUMBER):
                if v[p][7*i+d] == WORKING_DAY:
                    res = res+1
            if res == 1:
                end_res = end_res + 1
    return end_res


def evaluate_variant(v) -> bool:
    if not detect_if_valid_number_of_persons_per_weekend(v):
        return False
    if not detect_if_one_person_per_working_day(v):
        return False
    new_res = dict()
    new_res['cnt_1p'] = count_number_of_one_person_per_day_per_working_day(v)
    new_res['thread'] = v
    valid_res.append(new_res)
    return True


def detect_if_thread_respect_max_hours_smooth(th):
    m = MAX_WEEKLY_WORKING_HOURS / ONE_DAY_WORKING_HOUR
    for i in range(0, len(th)-7):
        cnt = th.count(WORKING_DAY, i, i+7)
        if cnt > m:
            print("Too many hours {}h detected at index: {}".format(cnt*12, i))
            return False
    return True


def get_shift_index(th) -> int:
    idx = 1
    for v in variant_list:
        if th == v:
            return idx
        idx = idx + 1
    raise ValueError("Shift index not found")


if __name__ == '__main__':
    ''' get the weeks thread base on an excel input '''
    wb = load_workbook("data/template_v1.xlsx")
    ws = wb.active
    a = ""
    for x in range(2, 2 + THREAD_TOTAL_WEEK_NUMBER):
        for y in range(2, 9):
            a += ws.cell(row=x, column=y).value
    print("Full thread is -> {}".format(a))

    if not detect_if_thread_respect_max_hours_smooth(a):
        exit(0)

    ''' get all possible week rotations
        ex: if we have the initial thread weekA-weekB-weekC
        we'll get:
          Th1: weekA-weekB-weekC
          Th2: weekB-weekC-weekA
          Th3: weekC-weekA-weekB
    '''
    size = len(a)
    for i in range(0, int(size/7)):
        r = a[7*i:] + a[:7*i]
        variant_list.append(r)
    # print(variant_list)

    ''' get all variants based on the number of person working
        ex: say we have 2 persons working and we have the variant_list defined above,
        we will have these variants:
        personA: Th1 personB: Th2
        personA: Th1 personB: Th3
        personA: Th2 personB: Th3
    '''
    all_variants = list(itertools.combinations(variant_list, PERSON_NUMBER))

    ''' for each variant, detect if the variant is interesting '''
    num_var = 0
    for i in range(len(all_variants)):
        # print("\nvariant {} -> {}".format(i, all_variants[i]))
        if evaluate_variant(all_variants[i]):
            num_var = num_var + 1

    print("Number of possible variants: {} out of {}".format(num_var, len(all_variants)))
    new_list = sorted(valid_res, key=lambda d: d['cnt_1p'])

    ''' render results '''
    top = new_list[0]
    print("\nThe best shift is: {}".format(top))
    for th in top['thread']:
        print("P: shift {} -> \t{}".format(get_shift_index(th), th))

    count_number_of_one_person_per_day_per_working_day(top['thread'])
    print("Number of days with 1 person working: {} / {}".format(top['cnt_1p'], THREAD_TOTAL_WEEK_NUMBER*7))
